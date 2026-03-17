"""Tests for corpus_aggregator — cross-directory report aggregation."""

import json
import sys
from pathlib import Path
from unittest import mock

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))
from corpus_aggregator import (
    _repo_key_from_dir,
    _dir_priority,
    _get_vertical,
    _get_arch_pattern,
    _get_business_domain,
    _get_actionable_findings,
    _finding_signature,
    _finding_deps,
    _all_deps,
    _needs_project_classification,
    _needs_finding_classification,
    _extract_error_type,
    load_reports,
    aggregate,
    reclassify_reports,
    generate_xlsx,
    print_summary,
    main,
)


# ── Fixtures ──


def _report(
    vertical="web_app",
    architectural_pattern="fastapi",
    business_domain="general",
    findings=None,
    deps=None,
):
    """Build a minimal mycode report dict."""
    return {
        "project": {
            "name": "test",
            "path": "/tmp/repo",
            "language": "python",
            "dependencies": [
                {"name": d, "installed_version": "1.0.0"}
                for d in (deps or [])
            ],
        },
        "vertical": vertical,
        "architectural_pattern": architectural_pattern,
        "business_domain": business_domain,
        "findings": findings or [],
        "statistics": {"scenarios_run": 5, "scenarios_passed": 3},
    }


def _finding(severity="warning", category="memory", title="High memory", deps=None):
    return {
        "severity": severity,
        "category": category,
        "title": title,
        "affected_dependencies": deps or [],
    }


def _write_report(results_dir, repo_name, report):
    """Write a report JSON into results_dir/repo_name/mycode-report.json."""
    repo_dir = results_dir / repo_name
    repo_dir.mkdir(parents=True, exist_ok=True)
    (repo_dir / "mycode-report.json").write_text(json.dumps(report))


# ── Unit Tests: Helpers ──


class TestRepoKey:
    def test_normalises_case(self):
        assert _repo_key_from_dir("Owner__Repo") == "owner__repo"

    def test_strips_whitespace(self):
        assert _repo_key_from_dir("  foo__bar  ") == "foo__bar"


class TestDirPriority:
    def test_results_highest(self):
        assert _dir_priority(Path("results")) == 3

    def test_corpus_results_lowest(self):
        assert _dir_priority(Path("corpus_results")) == 0

    def test_unknown_dir_defaults_zero(self):
        assert _dir_priority(Path("other")) == 0

    def test_priority_ordering(self):
        dirs = ["corpus_results", "corpus_results_timeout", "corpus_results_retry", "results"]
        priorities = [_dir_priority(Path(d)) for d in dirs]
        assert priorities == sorted(priorities)


class TestGetVertical:
    def test_returns_vertical(self):
        assert _get_vertical({"vertical": "ml_model"}) == "ml_model"

    def test_missing_defaults_unclassified(self):
        assert _get_vertical({}) == "unclassified"

    def test_none_defaults_unclassified(self):
        assert _get_vertical({"vertical": None}) == "unclassified"

    def test_empty_string_defaults_unclassified(self):
        assert _get_vertical({"vertical": ""}) == "unclassified"


class TestGetArchPattern:
    def test_returns_pattern(self):
        assert _get_arch_pattern({"architectural_pattern": "streamlit"}) == "streamlit"

    def test_missing_defaults_unclassified(self):
        assert _get_arch_pattern({}) == "unclassified"


class TestGetBusinessDomain:
    def test_returns_domain(self):
        assert _get_business_domain({"business_domain": "fintech"}) == "fintech"

    def test_missing_defaults_general(self):
        assert _get_business_domain({}) == "general"

    def test_none_defaults_general(self):
        assert _get_business_domain({"business_domain": None}) == "general"

    def test_empty_string_defaults_general(self):
        assert _get_business_domain({"business_domain": ""}) == "general"


class TestActionableFindings:
    def test_filters_critical_and_warning(self):
        findings = [
            _finding(severity="critical"),
            _finding(severity="warning"),
            _finding(severity="info"),
        ]
        result = _get_actionable_findings({"findings": findings})
        assert len(result) == 2
        assert all(f["severity"] in ("critical", "warning") for f in result)

    def test_empty_findings(self):
        assert _get_actionable_findings({"findings": []}) == []

    def test_no_findings_key(self):
        assert _get_actionable_findings({}) == []


class TestFindingSignature:
    def test_format(self):
        f = _finding(severity="critical", category="memory", title="OOM")
        assert _finding_signature(f) == "critical:memory:OOM"

    def test_truncates_at_120(self):
        f = _finding(title="x" * 200)
        assert len(_finding_signature(f)) == 120


class TestFindingDeps:
    def test_extracts_deps(self):
        f = _finding(deps=["pandas==1.0", "numpy"])
        assert _finding_deps(f) == ["pandas", "numpy"]

    def test_empty(self):
        assert _finding_deps(_finding()) == []


class TestAllDeps:
    def test_extracts_from_project(self):
        r = _report(deps=["flask", "sqlalchemy"])
        assert _all_deps(r) == ["flask", "sqlalchemy"]

    def test_empty_deps(self):
        assert _all_deps(_report()) == []


# ── Integration Tests: load_reports ──


class TestLoadReports:
    def test_loads_single_dir(self, tmp_path):
        results = tmp_path / "results"
        _write_report(results, "user__app", _report())

        reports = load_reports([results])
        assert len(reports) == 1
        assert "user__app" in reports

    def test_skips_missing_dir(self, tmp_path):
        reports = load_reports([tmp_path / "nonexistent"])
        assert reports == {}

    def test_skips_dir_without_report(self, tmp_path):
        results = tmp_path / "results"
        repo_dir = results / "user__app"
        repo_dir.mkdir(parents=True)
        # No mycode-report.json

        reports = load_reports([results])
        assert reports == {}

    def test_skips_invalid_json(self, tmp_path):
        results = tmp_path / "results"
        repo_dir = results / "user__app"
        repo_dir.mkdir(parents=True)
        (repo_dir / "mycode-report.json").write_text("not json{{{")

        reports = load_reports([results])
        assert reports == {}

    def test_dedup_higher_priority_wins(self, tmp_path):
        old_dir = tmp_path / "corpus_results"
        new_dir = tmp_path / "results"

        old_report = _report(vertical="old_vertical")
        new_report = _report(vertical="new_vertical")

        _write_report(old_dir, "user__app", old_report)
        _write_report(new_dir, "user__app", new_report)

        reports = load_reports([old_dir, new_dir])
        assert len(reports) == 1
        assert reports["user__app"]["vertical"] == "new_vertical"

    def test_dedup_order_independent(self, tmp_path):
        """Higher-priority dir wins regardless of scan order."""
        old_dir = tmp_path / "corpus_results"
        new_dir = tmp_path / "results"

        _write_report(old_dir, "user__app", _report(vertical="old"))
        _write_report(new_dir, "user__app", _report(vertical="new"))

        # Pass new_dir first, old_dir second — results/ should still win
        reports = load_reports([new_dir, old_dir])
        assert reports["user__app"]["vertical"] == "new"

    def test_dedup_case_insensitive(self, tmp_path):
        dir1 = tmp_path / "results"
        dir2 = tmp_path / "corpus_results"

        _write_report(dir1, "User__App", _report(vertical="v1"))
        _write_report(dir2, "user__app", _report(vertical="v2"))

        reports = load_reports([dir1, dir2])
        assert len(reports) == 1
        # results/ has higher priority
        assert reports["user__app"]["vertical"] == "v1"

    def test_multiple_repos(self, tmp_path):
        results = tmp_path / "results"
        _write_report(results, "user__app1", _report(vertical="web_app"))
        _write_report(results, "user__app2", _report(vertical="ml_model"))

        reports = load_reports([results])
        assert len(reports) == 2


# ── Integration Tests: aggregate ──


class TestAggregate:
    def test_basic_counts(self):
        reports = {
            "a": _report(vertical="web_app", architectural_pattern="fastapi"),
            "b": _report(vertical="web_app", architectural_pattern="flask"),
            "c": _report(vertical="ml_model", architectural_pattern="fastapi"),
        }
        result = aggregate(reports)

        assert result["total_repos"] == 3
        assert result["verticals"]["web_app"]["repo_count"] == 2
        assert result["verticals"]["ml_model"]["repo_count"] == 1
        assert result["architectural_patterns"]["fastapi"]["repo_count"] == 2
        assert result["architectural_patterns"]["flask"]["repo_count"] == 1

    def test_failure_rate(self):
        reports = {
            "clean": _report(vertical="web_app", findings=[]),
            "dirty": _report(vertical="web_app", findings=[_finding(severity="critical")]),
        }
        result = aggregate(reports)

        v = result["verticals"]["web_app"]
        assert v["repo_count"] == 2
        assert v["repos_with_issues"] == 1
        assert v["repos_clean"] == 1
        assert v["failure_rate"] == 0.5

    def test_info_findings_dont_count_as_issues(self):
        reports = {
            "a": _report(vertical="web_app", findings=[_finding(severity="info")]),
        }
        result = aggregate(reports)
        assert result["verticals"]["web_app"]["repos_with_issues"] == 0

    def test_top_signatures(self):
        findings = [
            _finding(severity="warning", category="mem", title="leak"),
            _finding(severity="warning", category="mem", title="leak"),
            _finding(severity="critical", category="cpu", title="spike"),
        ]
        reports = {"a": _report(vertical="v", findings=findings)}
        result = aggregate(reports)

        sigs = result["verticals"]["v"]["top_failure_signatures"]
        assert len(sigs) == 2
        assert sigs[0]["signature"] == "warning:mem:leak"
        assert sigs[0]["count"] == 2

    def test_top_signatures_capped_at_5(self):
        findings = [
            _finding(severity="warning", title=f"sig-{i}")
            for i in range(10)
        ]
        reports = {"a": _report(vertical="v", findings=findings)}
        result = aggregate(reports)
        assert len(result["verticals"]["v"]["top_failure_signatures"]) == 5

    def test_top_dependencies(self):
        findings = [
            _finding(severity="warning", deps=["pandas", "numpy"]),
            _finding(severity="critical", deps=["pandas"]),
        ]
        reports = {"a": _report(vertical="v", findings=findings, deps=["pandas", "numpy", "flask"])}
        result = aggregate(reports)

        deps = result["verticals"]["v"]["top_dependencies"]
        assert deps[0]["dependency"] == "pandas"
        assert deps[0]["failure_count"] == 2

    def test_dependency_failure_rate(self):
        findings = [_finding(severity="warning", deps=["pandas"])]
        reports = {
            "a": _report(vertical="v", findings=findings, deps=["pandas", "numpy", "flask"]),
        }
        result = aggregate(reports)

        v = result["verticals"]["v"]
        assert v["total_dependencies_seen"] == 3
        assert v["dependencies_with_failures"] == 1
        assert abs(v["dependency_failure_rate"] - 0.333) < 0.01

    def test_cross_tabulation(self):
        reports = {
            "a": _report(vertical="web_app", architectural_pattern="fastapi"),
            "b": _report(vertical="web_app", architectural_pattern="flask"),
            "c": _report(vertical="web_app", architectural_pattern="fastapi"),
            "d": _report(vertical="ml_model", architectural_pattern="fastapi"),
        }
        result = aggregate(reports)

        cross = result["cross_tabulation"]
        assert cross["web_app"]["fastapi"] == 2
        assert cross["web_app"]["flask"] == 1
        assert cross["ml_model"]["fastapi"] == 1
        assert "flask" not in cross["ml_model"]

    def test_old_format_defaults_to_unclassified(self):
        """Reports without vertical/architectural_pattern use 'unclassified'."""
        old_report = {
            "project": {"name": "old", "dependencies": []},
            "findings": [_finding(severity="warning")],
        }
        reports = {"a": old_report}
        result = aggregate(reports)

        assert "unclassified" in result["verticals"]
        assert result["verticals"]["unclassified"]["repo_count"] == 1

    def test_empty_reports(self):
        result = aggregate({})
        assert result["total_repos"] == 0
        assert result["verticals"] == {}
        assert result["architectural_patterns"] == {}
        assert result["cross_tabulation"] == {}
        assert result["business_domains"] == {}
        assert result["business_domain_vertical_cross"] == {}

    def test_business_domain_counts(self):
        reports = {
            "a": _report(business_domain="fintech"),
            "b": _report(business_domain="fintech"),
            "c": _report(business_domain="healthcare"),
        }
        result = aggregate(reports)

        assert result["business_domains"]["fintech"]["repo_count"] == 2
        assert result["business_domains"]["healthcare"]["repo_count"] == 1

    def test_business_domain_failure_rate(self):
        reports = {
            "clean": _report(business_domain="fintech", findings=[]),
            "dirty": _report(business_domain="fintech", findings=[_finding(severity="critical")]),
        }
        result = aggregate(reports)

        bd = result["business_domains"]["fintech"]
        assert bd["repo_count"] == 2
        assert bd["repos_with_issues"] == 1
        assert bd["failure_rate"] == 0.5

    def test_business_domain_top_signatures(self):
        findings = [
            _finding(severity="warning", category="mem", title="leak"),
            _finding(severity="warning", category="mem", title="leak"),
        ]
        reports = {"a": _report(business_domain="fintech", findings=findings)}
        result = aggregate(reports)

        sigs = result["business_domains"]["fintech"]["top_failure_signatures"]
        assert len(sigs) == 1
        assert sigs[0]["count"] == 2

    def test_business_domain_vertical_cross(self):
        reports = {
            "a": _report(vertical="web_app", business_domain="fintech"),
            "b": _report(vertical="dashboard", business_domain="fintech"),
            "c": _report(vertical="web_app", business_domain="healthcare"),
        }
        result = aggregate(reports)

        cross = result["business_domain_vertical_cross"]
        assert cross["fintech"]["web_app"] == 1
        assert cross["fintech"]["dashboard"] == 1
        assert cross["healthcare"]["web_app"] == 1


# ── Print Summary (smoke test) ──


class TestPrintSummary:
    def test_does_not_crash(self, capsys):
        reports = {
            "a": _report(vertical="web_app", findings=[_finding(severity="critical", deps=["flask"])], deps=["flask"]),
            "b": _report(vertical="ml_model"),
        }
        result = aggregate(reports)
        print_summary(result)

        captured = capsys.readouterr()
        assert "CORPUS AGGREGATE" in captured.out
        assert "web_app" in captured.out
        assert "ml_model" in captured.out

    def test_empty_corpus(self, capsys):
        result = aggregate({})
        print_summary(result)
        captured = capsys.readouterr()
        assert "0 repos" in captured.out


# ── Reclassification Helpers ──


class TestNeedsProjectClassification:
    def test_missing_vertical(self):
        assert _needs_project_classification({}) is True

    def test_null_vertical(self):
        assert _needs_project_classification({"vertical": None}) is True

    def test_empty_vertical(self):
        assert _needs_project_classification({"vertical": ""}) is True

    def test_missing_arch_pattern(self):
        assert _needs_project_classification({"vertical": "web_app"}) is True

    def test_already_classified(self):
        assert _needs_project_classification({
            "vertical": "web_app",
            "architectural_pattern": "fastapi",
            "business_domain": "general",
        }) is False

    def test_missing_business_domain(self):
        assert _needs_project_classification({
            "vertical": "web_app",
            "architectural_pattern": "fastapi",
        }) is True


class TestNeedsFindingClassification:
    def test_missing_domain(self):
        assert _needs_finding_classification({}) is True

    def test_null_domain(self):
        assert _needs_finding_classification({"failure_domain": None}) is True

    def test_already_classified(self):
        assert _needs_finding_classification({"failure_domain": "resource_exhaustion"}) is False


class TestExtractErrorType:
    def test_extracts_memory_error(self):
        f = {"description": "Process crashed with MemoryError at line 42"}
        assert _extract_error_type(f) == "MemoryError"

    def test_extracts_from_title(self):
        f = {"title": "TypeError in handler", "description": ""}
        assert _extract_error_type(f) == "TypeError"

    def test_returns_empty_when_none(self):
        f = {"description": "Something went wrong"}
        assert _extract_error_type(f) == ""

    def test_handles_missing_fields(self):
        assert _extract_error_type({}) == ""


# ── Reclassification Integration ──


class TestReclassifyReports:
    def test_reclassifies_null_vertical(self, tmp_path):
        """Report with null vertical gets classified from dependencies."""
        old_report = {
            "project": {
                "name": "test",
                "path": "/tmp/repo",
                "language": "python",
                "files_analyzed": 5,
                "dependencies": [
                    {"name": "streamlit", "installed_version": "1.0.0"},
                    {"name": "pandas", "installed_version": "2.0.0"},
                ],
            },
            "vertical": None,
            "architectural_pattern": None,
            "findings": [],
        }
        results = tmp_path / "corpus_results"
        _write_report(results, "user__dashboard", old_report)

        summary = reclassify_reports([results])

        assert summary["project_classified"] == 1
        assert summary["files_updated"] == 1

        # Read back and verify classification was written
        updated = json.loads(
            (results / "user__dashboard" / "mycode-report.json").read_text()
        )
        assert updated["vertical"] == "dashboard"
        assert updated["architectural_pattern"] == "dashboard"

    def test_reclassifies_findings(self, tmp_path):
        """Findings missing failure_domain get classified."""
        old_report = {
            "project": {
                "name": "test",
                "dependencies": [],
            },
            "vertical": "web_app",
            "architectural_pattern": "web_app",
            "business_domain": "general",
            "findings": [
                {
                    "title": "Memory growth under load",
                    "severity": "warning",
                    "category": "memory_profiling",
                    "description": "MemoryError after 1000 iterations",
                    "affected_dependencies": [],
                },
            ],
        }
        results = tmp_path / "corpus_results"
        _write_report(results, "user__app", old_report)

        summary = reclassify_reports([results])

        assert summary["findings_classified"] == 1
        assert summary["project_classified"] == 0  # already has vertical

        updated = json.loads(
            (results / "user__app" / "mycode-report.json").read_text()
        )
        finding = updated["findings"][0]
        assert finding["failure_domain"] == "resource_exhaustion"
        assert finding["operational_trigger"] is not None

    def test_skips_already_classified(self, tmp_path):
        """Reports with existing vertical and classified findings are untouched."""
        classified_report = {
            "project": {"name": "test", "dependencies": []},
            "vertical": "web_app",
            "architectural_pattern": "fastapi",
            "business_domain": "general",
            "findings": [
                {
                    "title": "OOM",
                    "severity": "critical",
                    "category": "memory_profiling",
                    "description": "",
                    "failure_domain": "resource_exhaustion",
                    "failure_pattern": "large_payload_oom",
                    "operational_trigger": "sustained_load",
                    "affected_dependencies": [],
                },
            ],
        }
        results = tmp_path / "corpus_results"
        _write_report(results, "user__app", classified_report)

        summary = reclassify_reports([results])

        assert summary["files_updated"] == 0
        assert summary["project_classified"] == 0
        assert summary["findings_classified"] == 0

    def test_no_deps_defaults_to_utility(self, tmp_path):
        """Report with no recognisable dependencies gets 'utility'."""
        old_report = {
            "project": {"name": "test", "dependencies": []},
            "vertical": None,
            "architectural_pattern": None,
            "findings": [],
        }
        results = tmp_path / "corpus_results"
        _write_report(results, "user__empty", old_report)

        reclassify_reports([results])

        updated = json.loads(
            (results / "user__empty" / "mycode-report.json").read_text()
        )
        assert updated["vertical"] == "utility"
        assert updated["architectural_pattern"] == "utility"

    def test_processes_all_dirs(self, tmp_path):
        """Reclassify processes all directories, not deduplicated."""
        old_report = {
            "project": {
                "name": "test",
                "dependencies": [{"name": "flask", "installed_version": "3.0"}],
            },
            "vertical": None,
            "architectural_pattern": None,
            "findings": [],
        }
        dir1 = tmp_path / "corpus_results"
        dir2 = tmp_path / "corpus_results_retry"
        _write_report(dir1, "user__app", old_report)
        _write_report(dir2, "user__app", old_report)

        summary = reclassify_reports([dir1, dir2])

        assert summary["files_updated"] == 2
        assert summary["project_classified"] == 2

    def test_skips_missing_dir(self, tmp_path):
        summary = reclassify_reports([tmp_path / "nonexistent"])
        assert summary["total_scanned"] == 0

    def test_summary_counts(self, tmp_path):
        """Verify summary structure and vertical breakdown."""
        old_report = {
            "project": {
                "name": "test",
                "dependencies": [{"name": "torch", "installed_version": "2.0"}],
            },
            "vertical": None,
            "architectural_pattern": None,
            "findings": [
                {
                    "title": "Scaling issue",
                    "severity": "warning",
                    "category": "data_volume_scaling",
                    "description": "",
                    "affected_dependencies": [],
                },
            ],
        }
        results = tmp_path / "corpus_results"
        _write_report(results, "user__ml", old_report)

        summary = reclassify_reports([results])

        assert summary["total_scanned"] == 1
        assert summary["files_updated"] == 1
        assert summary["project_classified"] == 1
        assert summary["findings_classified"] == 1
        assert "ml_model" in summary["verticals_assigned"]


class TestReclassifyCLI:
    def test_reclassify_flag(self, tmp_path):
        old_report = {
            "project": {
                "name": "test",
                "dependencies": [{"name": "fastapi", "installed_version": "0.100"}],
            },
            "vertical": None,
            "architectural_pattern": None,
            "findings": [],
        }
        results = tmp_path / "corpus_results"
        _write_report(results, "user__api", old_report)
        output = tmp_path / "out.json"

        main(["--dirs", str(results), "--output", str(output), "--reclassify"])

        # Report should be reclassified on disk
        updated = json.loads(
            (results / "user__api" / "mycode-report.json").read_text()
        )
        assert updated["vertical"] is not None
        assert updated["vertical"] != ""

        # Aggregate should use the new classification
        agg = json.loads(output.read_text())
        assert "unclassified" not in agg["verticals"]


# ── XLSX Output ──


class TestGenerateXlsx:
    def _aggregate_fixture(self):
        """Build a non-trivial aggregate result for xlsx tests."""
        reports = {
            "a": _report(
                vertical="web_app",
                architectural_pattern="fastapi",
                findings=[
                    _finding(severity="critical", category="memory", title="OOM", deps=["pandas"]),
                    _finding(severity="warning", category="cpu", title="Spike", deps=["numpy"]),
                ],
                deps=["pandas", "numpy", "flask"],
            ),
            "b": _report(
                vertical="web_app",
                architectural_pattern="flask",
                findings=[_finding(severity="warning", category="io", title="Blocking")],
                deps=["flask"],
            ),
            "c": _report(
                vertical="ml_model",
                architectural_pattern="fastapi",
                findings=[],
                deps=["scikit-learn"],
            ),
        }
        return aggregate(reports)

    def test_generates_xlsx_file(self, tmp_path):
        openpyxl = pytest.importorskip("openpyxl")
        result = self._aggregate_fixture()
        xlsx_path = tmp_path / "test.xlsx"

        assert generate_xlsx(result, xlsx_path) is True
        assert xlsx_path.exists()

        wb = openpyxl.load_workbook(str(xlsx_path))
        assert len(wb.sheetnames) == 8
        assert wb.sheetnames == [
            "Repos Per Vertical",
            "Top Failure Signatures",
            "Top Dependencies",
            "Repos Per Pattern",
            "Cross-Tabulation",
            "Repos Per Domain",
            "Domain Failure Signatures",
            "Domain x Vertical",
        ]

    def test_repos_per_vertical_sheet(self, tmp_path):
        openpyxl = pytest.importorskip("openpyxl")
        result = self._aggregate_fixture()
        xlsx_path = tmp_path / "test.xlsx"
        generate_xlsx(result, xlsx_path)

        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb["Repos Per Vertical"]
        # Header row
        assert ws.cell(row=1, column=1).value == "Vertical"
        assert ws.cell(row=1, column=5).value == "Failure Rate"
        # web_app has 2 repos (sorted first by count desc)
        assert ws.cell(row=2, column=1).value == "web_app"
        assert ws.cell(row=2, column=2).value == 2
        # ml_model has 1
        assert ws.cell(row=3, column=1).value == "ml_model"
        assert ws.cell(row=3, column=2).value == 1

    def test_failure_signatures_sheet(self, tmp_path):
        openpyxl = pytest.importorskip("openpyxl")
        result = self._aggregate_fixture()
        xlsx_path = tmp_path / "test.xlsx"
        generate_xlsx(result, xlsx_path)

        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb["Top Failure Signatures"]
        assert ws.cell(row=1, column=1).value == "Vertical"
        assert ws.cell(row=1, column=3).value == "Severity"
        # web_app should have signatures
        assert ws.cell(row=2, column=1).value == "web_app"
        assert ws.cell(row=2, column=2).value == 1  # rank

    def test_top_dependencies_sheet(self, tmp_path):
        openpyxl = pytest.importorskip("openpyxl")
        result = self._aggregate_fixture()
        xlsx_path = tmp_path / "test.xlsx"
        generate_xlsx(result, xlsx_path)

        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb["Top Dependencies"]
        assert ws.cell(row=1, column=3).value == "Dependency"
        # At least one row of data
        assert ws.cell(row=2, column=3).value is not None

    def test_repos_per_pattern_sheet(self, tmp_path):
        openpyxl = pytest.importorskip("openpyxl")
        result = self._aggregate_fixture()
        xlsx_path = tmp_path / "test.xlsx"
        generate_xlsx(result, xlsx_path)

        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb["Repos Per Pattern"]
        assert ws.cell(row=1, column=1).value == "Architectural Pattern"
        # fastapi: 2, flask: 1
        patterns = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(2, 4)}
        assert patterns["fastapi"] == 2
        assert patterns["flask"] == 1

    def test_cross_tabulation_sheet(self, tmp_path):
        openpyxl = pytest.importorskip("openpyxl")
        result = self._aggregate_fixture()
        xlsx_path = tmp_path / "test.xlsx"
        generate_xlsx(result, xlsx_path)

        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb["Cross-Tabulation"]
        assert ws.cell(row=1, column=1).value == "Vertical"
        # Column headers should be pattern names
        col_headers = [ws.cell(row=1, column=c).value for c in range(2, ws.max_column + 1)]
        assert "fastapi" in col_headers
        assert "flask" in col_headers

    def test_empty_aggregate(self, tmp_path):
        openpyxl = pytest.importorskip("openpyxl")
        result = aggregate({})
        xlsx_path = tmp_path / "empty.xlsx"

        assert generate_xlsx(result, xlsx_path) is True
        assert xlsx_path.exists()

        wb = openpyxl.load_workbook(str(xlsx_path))
        # Empty aggregate has no business_domain data, so no domain sheets
        assert len(wb.sheetnames) == 5

    def test_returns_false_without_openpyxl(self):
        result = aggregate({})
        with mock.patch.dict("sys.modules", {"openpyxl": None}):
            assert generate_xlsx(result, Path("/tmp/nope.xlsx")) is False


# ── CLI ──


class TestCLI:
    def test_main_writes_output(self, tmp_path):
        results = tmp_path / "results"
        _write_report(results, "user__app", _report())
        output = tmp_path / "out.json"

        main(["--dirs", str(results), "--output", str(output)])

        assert output.exists()
        data = json.loads(output.read_text())
        assert data["total_repos"] == 1
        assert "verticals" in data

    def test_main_default_output(self, tmp_path, monkeypatch):
        results = tmp_path / "results"
        _write_report(results, "user__app", _report())
        monkeypatch.chdir(tmp_path)

        main(["--dirs", str(results)])

        default_output = tmp_path / "corpus_aggregate.json"
        assert default_output.exists()

    def test_main_no_reports(self, tmp_path, capsys):
        with pytest.raises(SystemExit) as exc_info:
            main(["--dirs", str(tmp_path / "empty")])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()
        assert "No reports found" in captured.out

    def test_main_xlsx_flag(self, tmp_path):
        pytest.importorskip("openpyxl")
        results = tmp_path / "results"
        _write_report(results, "user__app", _report(
            findings=[_finding(severity="critical", deps=["flask"])],
            deps=["flask"],
        ))
        output = tmp_path / "out.json"

        main(["--dirs", str(results), "--output", str(output), "--xlsx"])

        xlsx_path = tmp_path / "out.xlsx"
        assert xlsx_path.exists()
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path))
        assert len(wb.sheetnames) == 8

    def test_main_multiple_dirs(self, tmp_path):
        dir1 = tmp_path / "results"
        dir2 = tmp_path / "corpus_results"
        _write_report(dir1, "user__app1", _report(vertical="v1"))
        _write_report(dir2, "user__app2", _report(vertical="v2"))
        output = tmp_path / "out.json"

        main(["--dirs", f"{dir1},{dir2}", "--output", str(output)])

        data = json.loads(output.read_text())
        assert data["total_repos"] == 2
