#!/usr/bin/env python3
"""corpus_aggregator — aggregate myCode reports across the full corpus.

Reads per-repo JSON reports from multiple result directories, deduplicates
by repo URL, and produces vertical/architectural_pattern breakdowns with
failure rates, top signatures, and dependency analysis.

Internal pre-launch tool.  Not shipped to users.

Usage:
    python scripts/corpus_aggregator.py
    python scripts/corpus_aggregator.py --output my_aggregate.json
    python scripts/corpus_aggregator.py --dirs results,corpus_results
    python scripts/corpus_aggregator.py --verbose
    python scripts/corpus_aggregator.py --reclassify --dirs corpus_results

No third-party dependencies — stdlib only (openpyxl optional for --xlsx).
"""

import argparse
import json
import logging
import sys
from collections import Counter, defaultdict
from pathlib import Path

logger = logging.getLogger("corpus_aggregator")

# ── Constants ──

_DEFAULT_DIRS = [
    "results",
    "corpus_results",
    "corpus_results_retry",
    "corpus_results_timeout",
]

# Priority order for deduplication: higher index wins.
_DIR_PRIORITY = {
    "corpus_results": 0,
    "corpus_results_timeout": 1,
    "corpus_results_retry": 2,
    "results": 3,
}


# ── Report Loading ──


def _repo_key_from_dir(repo_dir_name: str) -> str:
    """Normalise a repo directory name (owner__repo) to a dedup key."""
    return repo_dir_name.strip().lower()


def _dir_priority(dir_path: Path) -> int:
    """Return priority for a results directory (higher = preferred)."""
    return _DIR_PRIORITY.get(dir_path.name, 0)


def load_reports(dirs: list[Path]) -> dict[str, dict]:
    """Load and deduplicate reports from multiple result directories.

    Returns {repo_key: report_dict} with the highest-priority copy kept
    when the same repo appears in multiple directories.
    """
    # Track (priority, report) per repo key
    best: dict[str, tuple[int, dict]] = {}

    for results_dir in dirs:
        if not results_dir.is_dir():
            logger.debug("Skipping missing directory: %s", results_dir)
            continue

        priority = _dir_priority(results_dir)

        for repo_dir in results_dir.iterdir():
            if not repo_dir.is_dir():
                continue

            report_path = repo_dir / "mycode-report.json"
            if not report_path.is_file():
                continue

            try:
                report = json.loads(report_path.read_text())
            except (json.JSONDecodeError, OSError) as exc:
                logger.warning("Could not read %s: %s", report_path, exc)
                continue

            key = _repo_key_from_dir(repo_dir.name)
            existing = best.get(key)
            if existing is None or priority > existing[0]:
                best[key] = (priority, report)

    return {key: report for key, (_, report) in best.items()}


# ── Extraction Helpers ──


def _get_vertical(report: dict) -> str:
    return report.get("vertical") or "unclassified"


def _get_arch_pattern(report: dict) -> str:
    return report.get("architectural_pattern") or "unclassified"


def _get_business_domain(report: dict) -> str:
    return report.get("business_domain") or "general"


def _get_actionable_findings(report: dict) -> list[dict]:
    """Return findings with severity critical or warning."""
    return [
        f for f in report.get("findings", [])
        if f.get("severity") in ("critical", "warning")
    ]


def _finding_signature(finding: dict) -> str:
    sev = finding.get("severity", "")
    cat = finding.get("category", "")
    title = finding.get("title", "")
    return f"{sev}:{cat}:{title}"[:120]


def _finding_deps(finding: dict) -> list[str]:
    return [
        d.split("==")[0]
        for d in finding.get("affected_dependencies", [])
        if isinstance(d, str)
    ]


def _all_deps(report: dict) -> list[str]:
    """All dependency names from the project section."""
    return [
        d["name"]
        for d in report.get("project", {}).get("dependencies", [])
        if isinstance(d, dict) and d.get("name")
    ]


# ── Aggregation ──


def aggregate(reports: dict[str, dict]) -> dict:
    """Build the full aggregate summary from deduplicated reports."""
    # Per-vertical accumulators
    v_repo_count: Counter = Counter()
    v_repos_with_issues: Counter = Counter()
    v_signatures: dict[str, Counter] = defaultdict(Counter)
    v_deps: dict[str, Counter] = defaultdict(Counter)
    v_deps_with_failures: dict[str, set] = defaultdict(set)
    v_all_deps: dict[str, set] = defaultdict(set)

    # Per-architectural-pattern
    ap_repo_count: Counter = Counter()

    # Per-business-domain accumulators
    bd_repo_count: Counter = Counter()
    bd_repos_with_issues: Counter = Counter()
    bd_signatures: dict[str, Counter] = defaultdict(Counter)
    bd_deps: dict[str, Counter] = defaultdict(Counter)
    bd_deps_with_failures: dict[str, set] = defaultdict(set)
    bd_all_deps: dict[str, set] = defaultdict(set)

    # Cross-tabulations
    cross: Counter = Counter()
    bd_cross: Counter = Counter()

    for _key, report in reports.items():
        vertical = _get_vertical(report)
        arch = _get_arch_pattern(report)
        domain = _get_business_domain(report)

        v_repo_count[vertical] += 1
        ap_repo_count[arch] += 1
        bd_repo_count[domain] += 1
        cross[(vertical, arch)] += 1
        bd_cross[(domain, vertical)] += 1

        actionable = _get_actionable_findings(report)
        if actionable:
            v_repos_with_issues[vertical] += 1
            bd_repos_with_issues[domain] += 1

        for finding in actionable:
            sig = _finding_signature(finding)
            v_signatures[vertical][sig] += 1
            bd_signatures[domain][sig] += 1
            for dep in _finding_deps(finding):
                v_deps[vertical][dep] += 1
                v_deps_with_failures[vertical].add(dep)
                bd_deps[domain][dep] += 1
                bd_deps_with_failures[domain].add(dep)

        for dep in _all_deps(report):
            v_all_deps[vertical].add(dep)
            bd_all_deps[domain].add(dep)

    # Build per-vertical summaries
    verticals: dict[str, dict] = {}
    for v in sorted(v_repo_count):
        total = v_repo_count[v]
        with_issues = v_repos_with_issues.get(v, 0)
        all_d = v_all_deps.get(v, set())
        failed_d = v_deps_with_failures.get(v, set())

        verticals[v] = {
            "repo_count": total,
            "repos_with_issues": with_issues,
            "repos_clean": total - with_issues,
            "failure_rate": round(with_issues / total, 3) if total else 0.0,
            "top_failure_signatures": [
                {"signature": sig, "count": count}
                for sig, count in v_signatures.get(v, Counter()).most_common(5)
            ],
            "top_dependencies": [
                {"dependency": dep, "failure_count": count}
                for dep, count in v_deps.get(v, Counter()).most_common(5)
            ],
            "dependency_failure_rate": (
                round(len(failed_d) / len(all_d), 3) if all_d else 0.0
            ),
            "total_dependencies_seen": len(all_d),
            "dependencies_with_failures": len(failed_d),
        }

    # Build per-architectural-pattern summaries
    arch_patterns: dict[str, dict] = {}
    for ap in sorted(ap_repo_count):
        arch_patterns[ap] = {"repo_count": ap_repo_count[ap]}

    # Build per-business-domain summaries
    business_domains: dict[str, dict] = {}
    for bd in sorted(bd_repo_count):
        total = bd_repo_count[bd]
        with_issues = bd_repos_with_issues.get(bd, 0)
        all_d = bd_all_deps.get(bd, set())
        failed_d = bd_deps_with_failures.get(bd, set())

        business_domains[bd] = {
            "repo_count": total,
            "repos_with_issues": with_issues,
            "repos_clean": total - with_issues,
            "failure_rate": round(with_issues / total, 3) if total else 0.0,
            "top_failure_signatures": [
                {"signature": sig, "count": count}
                for sig, count in bd_signatures.get(bd, Counter()).most_common(5)
            ],
            "top_dependencies": [
                {"dependency": dep, "failure_count": count}
                for dep, count in bd_deps.get(bd, Counter()).most_common(5)
            ],
            "dependency_failure_rate": (
                round(len(failed_d) / len(all_d), 3) if all_d else 0.0
            ),
            "total_dependencies_seen": len(all_d),
            "dependencies_with_failures": len(failed_d),
        }

    # Build cross-tabulation: vertical × architectural_pattern
    all_verticals = sorted(v_repo_count)
    all_patterns = sorted(ap_repo_count)
    cross_tab: dict[str, dict[str, int]] = {}
    for v in all_verticals:
        row: dict[str, int] = {}
        for ap in all_patterns:
            count = cross.get((v, ap), 0)
            if count > 0:
                row[ap] = count
        cross_tab[v] = row

    # Build cross-tabulation: business_domain × vertical
    all_domains = sorted(bd_repo_count)
    bd_vert_cross: dict[str, dict[str, int]] = {}
    for bd in all_domains:
        row: dict[str, int] = {}
        for v in all_verticals:
            count = bd_cross.get((bd, v), 0)
            if count > 0:
                row[v] = count
        bd_vert_cross[bd] = row

    return {
        "total_repos": len(reports),
        "verticals": verticals,
        "architectural_patterns": arch_patterns,
        "cross_tabulation": cross_tab,
        "business_domains": business_domains,
        "business_domain_vertical_cross": bd_vert_cross,
    }


# ── Terminal Output ──


def _print_table(headers: list[str], rows: list[list], col_widths: list[int] | None = None) -> None:
    """Print a simple aligned table."""
    if not col_widths:
        col_widths = [
            max(len(str(h)), *(len(str(r[i])) for r in rows) if rows else [len(str(h))])
            for i, h in enumerate(headers)
        ]

    header_line = "  ".join(str(h).ljust(w) for h, w in zip(headers, col_widths))
    print(header_line)
    print("  ".join("-" * w for w in col_widths))
    for row in rows:
        print("  ".join(str(c).ljust(w) for c, w in zip(row, col_widths)))


def print_summary(result: dict) -> None:
    """Print a human-readable summary to the terminal."""
    print(f"\n{'=' * 60}")
    print(f"  CORPUS AGGREGATE — {result['total_repos']} repos")
    print(f"{'=' * 60}\n")

    # ── Repos per vertical ──
    print("REPOS PER VERTICAL")
    rows = []
    for v, data in sorted(result["verticals"].items(), key=lambda x: -x[1]["repo_count"]):
        rows.append([
            v,
            data["repo_count"],
            f"{data['failure_rate'] * 100:.1f}%",
            data["repos_with_issues"],
            data["repos_clean"],
        ])
    _print_table(
        ["Vertical", "Repos", "Fail%", "Issues", "Clean"],
        rows,
        [30, 6, 7, 7, 6],
    )

    # ── Top signatures per vertical ──
    print(f"\n{'─' * 60}")
    print("TOP FAILURE SIGNATURES PER VERTICAL\n")
    for v, data in sorted(result["verticals"].items(), key=lambda x: -x[1]["repo_count"]):
        sigs = data["top_failure_signatures"]
        if not sigs:
            continue
        print(f"  {v} ({data['repo_count']} repos):")
        for entry in sigs:
            print(f"    {entry['count']:3d}  {entry['signature']}")
        print()

    # ── Top dependencies per vertical ──
    print(f"{'─' * 60}")
    print("TOP DEPENDENCIES WITH FAILURES PER VERTICAL\n")
    for v, data in sorted(result["verticals"].items(), key=lambda x: -x[1]["repo_count"]):
        deps = data["top_dependencies"]
        if not deps:
            continue
        print(f"  {v} ({data['repo_count']} repos, dep fail rate: {data['dependency_failure_rate'] * 100:.1f}%):")
        for entry in deps:
            print(f"    {entry['failure_count']:3d}  {entry['dependency']}")
        print()

    # ── Repos per architectural pattern ──
    print(f"{'─' * 60}")
    print("REPOS PER ARCHITECTURAL PATTERN")
    rows = []
    for ap, data in sorted(result["architectural_patterns"].items(), key=lambda x: -x[1]["repo_count"]):
        rows.append([ap, data["repo_count"]])
    _print_table(
        ["Pattern", "Repos"],
        rows,
        [30, 6],
    )

    # ── Repos per business domain ──
    if result.get("business_domains"):
        print(f"\n{'─' * 60}")
        print("REPOS PER BUSINESS DOMAIN")
        rows = []
        for bd, data in sorted(result["business_domains"].items(), key=lambda x: -x[1]["repo_count"]):
            rows.append([
                bd,
                data["repo_count"],
                f"{data['failure_rate'] * 100:.1f}%",
                data["repos_with_issues"],
                data["repos_clean"],
            ])
        _print_table(
            ["Domain", "Repos", "Fail%", "Issues", "Clean"],
            rows,
            [30, 6, 7, 7, 6],
        )

        # ── Top signatures per business domain ──
        print(f"\n{'─' * 60}")
        print("TOP FAILURE SIGNATURES PER BUSINESS DOMAIN\n")
        for bd, data in sorted(result["business_domains"].items(), key=lambda x: -x[1]["repo_count"]):
            sigs = data["top_failure_signatures"]
            if not sigs:
                continue
            print(f"  {bd} ({data['repo_count']} repos):")
            for entry in sigs:
                print(f"    {entry['count']:3d}  {entry['signature']}")
            print()

    # ── Business domain × vertical cross-tabulation ──
    bd_vert_cross = result.get("business_domain_vertical_cross", {})
    if bd_vert_cross:
        print(f"{'─' * 60}")
        print("CROSS-TABULATION: BUSINESS DOMAIN x VERTICAL\n")
        all_verts = sorted({v for row in bd_vert_cross.values() for v in row})
        if all_verts:
            vert_width = max(len(v) for v in all_verts)
            vert_width = max(vert_width, 5)
            dom_width = max(len(bd) for bd in bd_vert_cross) if bd_vert_cross else 10

            header = ["Domain".ljust(dom_width)] + [v[:vert_width].ljust(vert_width) for v in all_verts]
            print("  ".join(header))
            print("  ".join("-" * len(h) for h in header))
            for bd in sorted(bd_vert_cross):
                cells = [bd.ljust(dom_width)]
                for v in all_verts:
                    val = bd_vert_cross[bd].get(v, 0)
                    cells.append((str(val) if val else ".").ljust(vert_width))
                print("  ".join(cells))
        print()

    # ── Cross-tabulation ──
    cross = result["cross_tabulation"]
    if cross:
        print(f"\n{'─' * 60}")
        print("CROSS-TABULATION: VERTICAL x ARCHITECTURAL PATTERN\n")
        all_patterns = sorted({
            ap for row in cross.values() for ap in row
        })
        # Only show patterns that appear
        if all_patterns:
            pat_width = max(len(p) for p in all_patterns)
            pat_width = max(pat_width, 5)
            vert_width = max(len(v) for v in cross) if cross else 10

            header = ["Vertical".ljust(vert_width)] + [p[:pat_width].ljust(pat_width) for p in all_patterns]
            print("  ".join(header))
            print("  ".join("-" * len(h) for h in header))
            for v in sorted(cross):
                cells = [v.ljust(vert_width)]
                for ap in all_patterns:
                    val = cross[v].get(ap, 0)
                    cells.append((str(val) if val else ".").ljust(pat_width))
                print("  ".join(cells))

    print(f"\n{'=' * 60}\n")


# ── Reclassification ──


def _needs_project_classification(report: dict) -> bool:
    """True if the report is missing vertical, architectural_pattern, or business_domain."""
    return (
        not report.get("vertical")
        or not report.get("architectural_pattern")
        or not report.get("business_domain")
    )


def _needs_finding_classification(finding: dict) -> bool:
    """True if a finding is missing failure_domain classification."""
    return not finding.get("failure_domain")


def _extract_error_type(finding: dict) -> str:
    """Best-effort extraction of error type from finding text."""
    desc = finding.get("description", "") or ""
    title = finding.get("title", "") or ""
    combined = f"{title} {desc}"
    # Look for common Python/JS error type names
    import re
    match = re.search(
        r'\b(MemoryError|TimeoutError|TypeError|ValueError|KeyError|IndexError'
        r'|ImportError|ModuleNotFoundError|ConnectionError|ConnectionRefusedError'
        r'|AttributeError|RuntimeError|OSError|FileNotFoundError|PermissionError'
        r'|UnicodeDecodeError|JSONDecodeError|BrokenPipeError)\b',
        combined,
    )
    return match.group(1) if match else ""


def reclassify_reports(dirs: list[Path]) -> dict:
    """Reclassify unclassified reports in place.

    Reads each mycode-report.json, applies project-level and finding-level
    classifiers where fields are missing, and writes updated JSON back.

    Unlike load_reports(), this processes ALL copies (not deduplicated) since
    each file on disk needs to be updated independently.

    Returns a summary dict with counts.
    """
    # Import myCode classifiers
    try:
        from mycode.classifiers import classify_project, classify_finding
    except ImportError:
        logger.error(
            "Cannot import mycode.classifiers — is mycode installed? "
            "Run: pip install -e ."
        )
        return {"error": "mycode not installed", "reports_updated": 0}

    total_scanned = 0
    project_classified = 0
    findings_classified = 0
    files_updated = 0
    vertical_counts: Counter = Counter()

    for results_dir in dirs:
        if not results_dir.is_dir():
            logger.debug("Skipping missing directory: %s", results_dir)
            continue

        for repo_dir in sorted(results_dir.iterdir()):
            if not repo_dir.is_dir():
                continue

            report_path = repo_dir / "mycode-report.json"
            if not report_path.is_file():
                continue

            try:
                report = json.loads(report_path.read_text())
            except (json.JSONDecodeError, OSError) as exc:
                logger.warning("Could not read %s: %s", report_path, exc)
                continue

            total_scanned += 1
            modified = False

            # Project-level classification
            if _needs_project_classification(report):
                deps = _all_deps(report)
                file_count = report.get("project", {}).get("files_analyzed", 0)
                project_name = report.get("project", {}).get("name", "")
                cls = classify_project(
                    dependencies=deps,
                    file_count=file_count,
                    project_name=project_name or "",
                )
                report["vertical"] = cls["vertical"]
                report["architectural_pattern"] = cls["architectural_pattern"]
                report["business_domain"] = cls["business_domain"]
                project_classified += 1
                modified = True
                vertical_counts[cls["vertical"]] += 1

            # Finding-level classification
            for finding in report.get("findings", []):
                if _needs_finding_classification(finding):
                    error_type = _extract_error_type(finding)
                    cls = classify_finding(
                        scenario_name=finding.get("title", ""),
                        scenario_category=finding.get("category", ""),
                        error_type=error_type,
                        error_details=finding.get("description", ""),
                    )
                    finding["failure_domain"] = cls["failure_domain"]
                    finding["failure_pattern"] = cls["failure_pattern"]
                    finding["operational_trigger"] = cls["operational_trigger"]
                    findings_classified += 1
                    modified = True

            if modified:
                report_path.write_text(json.dumps(report, indent=2) + "\n")
                files_updated += 1
                logger.debug("  Updated %s", report_path)

    summary = {
        "total_scanned": total_scanned,
        "files_updated": files_updated,
        "project_classified": project_classified,
        "findings_classified": findings_classified,
        "verticals_assigned": dict(vertical_counts.most_common()),
    }

    logger.info(
        "Reclassification complete: %d/%d reports updated "
        "(%d project, %d findings)",
        files_updated, total_scanned, project_classified, findings_classified,
    )

    return summary


# ── XLSX Output ──


def generate_xlsx(result: dict, xlsx_path: Path) -> bool:
    """Generate a formatted xlsx report from aggregate results.

    Requires openpyxl.  Returns True on success, False if openpyxl missing.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning(
            "openpyxl not installed — skipping xlsx report. "
            "Install with: pip install openpyxl"
        )
        return False

    # ── Styles ──
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    wb = Workbook()

    # ── Sheet 1: Repos Per Vertical ──
    ws1 = wb.active
    ws1.title = "Repos Per Vertical"
    write_header(ws1, ["Vertical", "Repos", "With Issues", "Clean", "Failure Rate",
                       "Deps Seen", "Deps Failed", "Dep Fail Rate"])

    for i, (v, data) in enumerate(
        sorted(result["verticals"].items(), key=lambda x: -x[1]["repo_count"]), 2
    ):
        ws1.cell(row=i, column=1, value=v)
        ws1.cell(row=i, column=2, value=data["repo_count"])
        ws1.cell(row=i, column=3, value=data["repos_with_issues"])
        ws1.cell(row=i, column=4, value=data["repos_clean"])
        ws1.cell(row=i, column=5, value=f"{data['failure_rate'] * 100:.1f}%")
        ws1.cell(row=i, column=6, value=data["total_dependencies_seen"])
        ws1.cell(row=i, column=7, value=data["dependencies_with_failures"])
        ws1.cell(row=i, column=8, value=f"{data['dependency_failure_rate'] * 100:.1f}%")

    auto_width(ws1)

    # ── Sheet 2: Top Failure Signatures ──
    ws2 = wb.create_sheet("Top Failure Signatures")
    write_header(ws2, ["Vertical", "Rank", "Severity", "Category", "Title", "Count"])

    row = 2
    for v, data in sorted(result["verticals"].items(), key=lambda x: -x[1]["repo_count"]):
        for rank, entry in enumerate(data["top_failure_signatures"], 1):
            raw_sig = entry["signature"]
            parts = raw_sig.split(":", 2)
            severity = parts[0] if len(parts) >= 1 else ""
            category = parts[1] if len(parts) >= 2 else ""
            title = parts[2] if len(parts) >= 3 else raw_sig

            ws2.cell(row=row, column=1, value=v)
            ws2.cell(row=row, column=2, value=rank)
            ws2.cell(row=row, column=3, value=severity)
            ws2.cell(row=row, column=4, value=category)
            ws2.cell(row=row, column=5, value=title)
            ws2.cell(row=row, column=6, value=entry["count"])
            row += 1

    auto_width(ws2)

    # ── Sheet 3: Top Dependencies ──
    ws3 = wb.create_sheet("Top Dependencies")
    write_header(ws3, ["Vertical", "Rank", "Dependency", "Failure Count"])

    row = 2
    for v, data in sorted(result["verticals"].items(), key=lambda x: -x[1]["repo_count"]):
        for rank, entry in enumerate(data["top_dependencies"], 1):
            ws3.cell(row=row, column=1, value=v)
            ws3.cell(row=row, column=2, value=rank)
            ws3.cell(row=row, column=3, value=entry["dependency"])
            ws3.cell(row=row, column=4, value=entry["failure_count"])
            row += 1

    auto_width(ws3)

    # ── Sheet 4: Repos Per Pattern ──
    ws4 = wb.create_sheet("Repos Per Pattern")
    write_header(ws4, ["Architectural Pattern", "Repos"])

    for i, (ap, data) in enumerate(
        sorted(result["architectural_patterns"].items(), key=lambda x: -x[1]["repo_count"]), 2
    ):
        ws4.cell(row=i, column=1, value=ap)
        ws4.cell(row=i, column=2, value=data["repo_count"])

    auto_width(ws4)

    # ── Sheet 5: Cross-Tabulation ──
    cross = result["cross_tabulation"]
    all_patterns = sorted({ap for row_data in cross.values() for ap in row_data})

    ws5 = wb.create_sheet("Cross-Tabulation")
    write_header(ws5, ["Vertical"] + all_patterns)

    for i, v in enumerate(sorted(cross), 2):
        ws5.cell(row=i, column=1, value=v)
        for j, ap in enumerate(all_patterns, 2):
            val = cross[v].get(ap, 0)
            ws5.cell(row=i, column=j, value=val if val else None)

    auto_width(ws5)

    # ── Sheet 6: Repos Per Business Domain ──
    if result.get("business_domains"):
        ws6 = wb.create_sheet("Repos Per Domain")
        write_header(ws6, ["Business Domain", "Repos", "With Issues", "Clean", "Failure Rate",
                           "Deps Seen", "Deps Failed", "Dep Fail Rate"])

        for i, (bd, data) in enumerate(
            sorted(result["business_domains"].items(), key=lambda x: -x[1]["repo_count"]), 2
        ):
            ws6.cell(row=i, column=1, value=bd)
            ws6.cell(row=i, column=2, value=data["repo_count"])
            ws6.cell(row=i, column=3, value=data["repos_with_issues"])
            ws6.cell(row=i, column=4, value=data["repos_clean"])
            ws6.cell(row=i, column=5, value=f"{data['failure_rate'] * 100:.1f}%")
            ws6.cell(row=i, column=6, value=data["total_dependencies_seen"])
            ws6.cell(row=i, column=7, value=data["dependencies_with_failures"])
            ws6.cell(row=i, column=8, value=f"{data['dependency_failure_rate'] * 100:.1f}%")

        auto_width(ws6)

    # ── Sheet 7: Domain Failure Signatures ──
    if result.get("business_domains"):
        ws7 = wb.create_sheet("Domain Failure Signatures")
        write_header(ws7, ["Business Domain", "Rank", "Severity", "Category", "Title", "Count"])

        row = 2
        for bd, data in sorted(result["business_domains"].items(), key=lambda x: -x[1]["repo_count"]):
            for rank, entry in enumerate(data["top_failure_signatures"], 1):
                raw_sig = entry["signature"]
                parts = raw_sig.split(":", 2)
                severity = parts[0] if len(parts) >= 1 else ""
                category = parts[1] if len(parts) >= 2 else ""
                title = parts[2] if len(parts) >= 3 else raw_sig

                ws7.cell(row=row, column=1, value=bd)
                ws7.cell(row=row, column=2, value=rank)
                ws7.cell(row=row, column=3, value=severity)
                ws7.cell(row=row, column=4, value=category)
                ws7.cell(row=row, column=5, value=title)
                ws7.cell(row=row, column=6, value=entry["count"])
                row += 1

        auto_width(ws7)

    # ── Sheet 8: Business Domain × Vertical ──
    bd_vert_cross = result.get("business_domain_vertical_cross", {})
    if bd_vert_cross:
        all_verts_bd = sorted({v for row_data in bd_vert_cross.values() for v in row_data})

        ws8 = wb.create_sheet("Domain x Vertical")
        write_header(ws8, ["Business Domain"] + all_verts_bd)

        for i, bd in enumerate(sorted(bd_vert_cross), 2):
            ws8.cell(row=i, column=1, value=bd)
            for j, v in enumerate(all_verts_bd, 2):
                val = bd_vert_cross[bd].get(v, 0)
                ws8.cell(row=i, column=j, value=val if val else None)

        auto_width(ws8)

    # ── Save ──
    wb.save(str(xlsx_path))
    logger.info("Wrote xlsx report to %s", xlsx_path)
    return True


# ── CLI ──


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Aggregate myCode reports across the full corpus.",
    )
    parser.add_argument(
        "--output", "-o",
        default="corpus_aggregate.json",
        help="Output JSON file path (default: corpus_aggregate.json)",
    )
    parser.add_argument(
        "--dirs",
        default=None,
        help="Comma-separated list of result directories to scan (default: results,corpus_results,corpus_results_retry,corpus_results_timeout)",
    )
    parser.add_argument(
        "--reclassify",
        action="store_true",
        help="Reclassify unclassified reports in place before aggregating (requires mycode installed)",
    )
    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="Generate corpus_aggregate.xlsx alongside JSON (requires openpyxl)",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable debug logging",
    )

    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    if args.dirs:
        dirs = [Path(d.strip()) for d in args.dirs.split(",")]
    else:
        dirs = [Path(d) for d in _DEFAULT_DIRS]

    logger.info("Scanning directories: %s", ", ".join(str(d) for d in dirs))

    if args.reclassify:
        reclass_summary = reclassify_reports(dirs)
        updated = reclass_summary.get("files_updated", 0)
        proj = reclass_summary.get("project_classified", 0)
        fnd = reclass_summary.get("findings_classified", 0)
        print(f"\nReclassified {updated} reports ({proj} project, {fnd} findings)")
        verticals = reclass_summary.get("verticals_assigned", {})
        if verticals:
            print("Verticals assigned:")
            for v, count in sorted(verticals.items(), key=lambda x: -x[1]):
                print(f"  {count:4d}  {v}")
            print()

    reports = load_reports(dirs)
    logger.info("Loaded %d deduplicated reports", len(reports))

    if not reports:
        print("No reports found.")
        sys.exit(0)

    result = aggregate(reports)

    output_path = Path(args.output)
    output_path.write_text(json.dumps(result, indent=2) + "\n")
    logger.info("Wrote aggregate to %s", output_path)

    print_summary(result)
    print(f"JSON written to {output_path}")

    if args.xlsx:
        xlsx_path = output_path.with_suffix(".xlsx")
        if generate_xlsx(result, xlsx_path):
            print(f"XLSX written to {xlsx_path}")


if __name__ == "__main__":
    main()
