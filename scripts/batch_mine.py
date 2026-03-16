#!/usr/bin/env python3
"""batch_mine — run myCode against repos discovered by repo_hunter.

Reads discovered_repos.json, clones each repo to a temp directory,
runs ``mycode --offline --non-interactive --json-output --skip-version-check``,
and collects the JSON report + discovery files.  Produces a per-repo
results directory and an aggregate summary (batch_results.json).

Internal pre-launch tool.  Not shipped to users.

Usage:
    python scripts/batch_mine.py                              # defaults
    python scripts/batch_mine.py --input repos.json           # custom input
    python scripts/batch_mine.py --max-repos 10               # limit
    python scripts/batch_mine.py --results-dir ./my_results   # custom output
    python scripts/batch_mine.py --timeout 900                # 15-min timeout for slow repos
    python scripts/batch_mine.py --report                     # also generate batch_report.xlsx
    python scripts/batch_mine.py --report-only                # regenerate reports from existing results

No third-party dependencies — stdlib only (openpyxl optional for --report).
"""

import argparse
import gc
import json
import logging
import os
import signal
import shutil
import subprocess
import sys
import tempfile
import time
from collections import Counter
from pathlib import Path

logger = logging.getLogger("batch_mine")

_IS_WINDOWS = sys.platform == "win32"

# ── Constants ──

_DISCOVERIES_DIR = Path.home() / ".mycode" / "discoveries"
_CLONE_TIMEOUT = 120          # seconds
_DEFAULT_MYCODE_TIMEOUT = 300  # seconds
_PROCESSED_REPOS_PATH = Path.home() / ".mycode" / "processed_repos.txt"


# ── Dedup helpers ──


def _load_processed_repos() -> set[str]:
    """Read the processed-repos file into a set of normalised URLs."""
    if not _PROCESSED_REPOS_PATH.is_file():
        return set()
    try:
        text = _PROCESSED_REPOS_PATH.read_text(encoding="utf-8")
    except OSError as exc:
        logger.warning("Could not read %s: %s", _PROCESSED_REPOS_PATH, exc)
        return set()
    return {line.strip().rstrip("/") for line in text.splitlines() if line.strip()}


def _record_processed_repo(repo_url: str) -> None:
    """Append a successfully-processed repo URL to the dedup file."""
    _PROCESSED_REPOS_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(_PROCESSED_REPOS_PATH, "a", encoding="utf-8") as fh:
            fh.write(repo_url.strip().rstrip("/") + "\n")
            fh.flush()
    except OSError as exc:
        logger.warning("Could not write to %s: %s", _PROCESSED_REPOS_PATH, exc)


# ── Helpers ──


def _kill_process_group(proc: subprocess.Popen) -> None:
    """Kill the entire process tree rooted at *proc*.

    Mirrors session.py's ``_kill_process_tree``.

    On POSIX: we launch children with ``start_new_session=True`` so
    ``os.killpg`` kills the entire process group (SIGTERM then SIGKILL).

    On Windows: we launch children with ``CREATE_NEW_PROCESS_GROUP`` and
    use ``taskkill /T /F /PID`` to kill the entire process tree.
    """
    if proc is None or proc.poll() is not None:
        return

    if _IS_WINDOWS:
        _kill_process_group_win32(proc)
    else:
        _kill_process_group_posix(proc)


def _kill_process_group_posix(proc: subprocess.Popen) -> None:
    """POSIX: SIGTERM the process group, then SIGKILL if needed."""
    pgid = None
    try:
        pgid = os.getpgid(proc.pid)
    except OSError:
        pass

    # 1. SIGTERM the group (graceful — lets mycode's cleanup handlers run).
    #    mycode's SIGTERM handler calls _kill_process_tree on harness children,
    #    which itself takes up to ~6s (SIGTERM 3s + SIGKILL 3s).  Give 10s grace
    #    so mycode can finish cleaning up before we escalate to SIGKILL.
    if pgid is not None and pgid > 0:
        try:
            os.killpg(pgid, signal.SIGTERM)
        except OSError:
            pass
    else:
        try:
            proc.terminate()
        except OSError:
            pass

    try:
        proc.wait(timeout=10)
        return
    except subprocess.TimeoutExpired:
        pass

    # 2. SIGKILL the group (forceful)
    if pgid is not None and pgid > 0:
        try:
            os.killpg(pgid, signal.SIGKILL)
        except OSError:
            pass
    else:
        try:
            proc.kill()
        except OSError:
            pass

    try:
        proc.wait(timeout=5)
    except subprocess.TimeoutExpired:
        logger.warning("Process %d did not exit after SIGKILL", proc.pid)


def _kill_process_group_win32(proc: subprocess.Popen) -> None:
    """Windows: use taskkill /T /F to kill the entire process tree."""
    # 1. Graceful: terminate the root process
    try:
        proc.terminate()
    except OSError:
        pass

    try:
        proc.wait(timeout=10)
        return
    except subprocess.TimeoutExpired:
        pass

    # 2. Forceful: taskkill /T (tree) /F (force) /PID
    try:
        subprocess.run(
            ["taskkill", "/T", "/F", "/PID", str(proc.pid)],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=10,
        )
    except (OSError, subprocess.TimeoutExpired):
        pass

    try:
        proc.wait(timeout=5)
    except subprocess.TimeoutExpired:
        logger.warning(
            "Process %d did not exit after taskkill /T /F", proc.pid
        )


def _clone_repo(clone_url: str, dest: Path) -> bool:
    """Shallow-clone a repo. Returns True on success."""
    proc = None
    try:
        popen_kwargs: dict = {
            "stdout": subprocess.PIPE,
            "stderr": subprocess.PIPE,
        }
        if _IS_WINDOWS:
            popen_kwargs["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP
        else:
            popen_kwargs["start_new_session"] = True
        proc = subprocess.Popen(
            ["git", "clone", "--depth", "1", "--single-branch", clone_url, str(dest)],
            **popen_kwargs,
        )
        _, stderr = proc.communicate(timeout=_CLONE_TIMEOUT)
        if proc.returncode != 0:
            logger.warning(
                "Clone failed: %s — %s",
                clone_url,
                stderr.decode(errors="replace")[:200],
            )
            return False
        return True
    except subprocess.TimeoutExpired:
        _kill_process_group(proc)
        logger.warning("Clone timed out after %ds: %s", _CLONE_TIMEOUT, clone_url)
        return False
    except OSError as exc:
        if proc is not None:
            _kill_process_group(proc)
        logger.warning("Clone failed: %s — %s", clone_url, str(exc)[:200])
        return False


def _snapshot_discoveries() -> set[str]:
    """Return the set of discovery file names currently in ~/.mycode/discoveries/."""
    if not _DISCOVERIES_DIR.is_dir():
        return set()
    return {f.name for f in _DISCOVERIES_DIR.iterdir() if f.suffix == ".json"}


def _collect_new_discoveries(before: set[str]) -> list[Path]:
    """Return paths to discovery files created after *before* snapshot."""
    if not _DISCOVERIES_DIR.is_dir():
        return []
    return [
        f for f in _DISCOVERIES_DIR.iterdir()
        if f.suffix == ".json" and f.name not in before
    ]


def _run_mycode(project_path: Path, timeout: int = _DEFAULT_MYCODE_TIMEOUT) -> tuple[int, str, str]:
    """Run mycode CLI against a project. Returns (returncode, stdout, stderr).

    On POSIX: uses ``start_new_session=True`` so that mycode and ALL its child
    processes share a single process group.  On timeout we kill the entire group
    with ``os.killpg``.

    On Windows: uses ``CREATE_NEW_PROCESS_GROUP`` and ``taskkill /T /F`` for
    tree-wide termination on timeout.
    """
    cmd = [
        sys.executable, "-m", "mycode",
        str(project_path),
        "--offline",
        "--non-interactive",
        "--json-output",
        "--skip-version-check",
    ]
    proc = None
    try:
        popen_kwargs: dict = {
            "stdout": subprocess.PIPE,
            "stderr": subprocess.PIPE,
            "text": True,
        }
        if _IS_WINDOWS:
            popen_kwargs["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP
        else:
            popen_kwargs["start_new_session"] = True
        proc = subprocess.Popen(cmd, **popen_kwargs)
        stdout, stderr = proc.communicate(timeout=timeout)
        return proc.returncode, stdout, stderr
    except subprocess.TimeoutExpired:
        if proc is not None:
            _kill_process_group(proc)
            # Drain any remaining buffered pipe data after killing the group
            try:
                stdout, stderr = proc.communicate(timeout=5)
            except (subprocess.TimeoutExpired, OSError, ValueError):
                stdout, stderr = "", ""
        else:
            stdout, stderr = "", ""
        return -1, stdout or "", stderr + "\nmycode timed out" if stderr else "mycode timed out"
    except Exception as exc:
        if proc is not None:
            _kill_process_group(proc)
        return -1, "", str(exc)


def _safe_name(repo_url: str) -> str:
    """Derive a filesystem-safe name from a repo URL.

    ``https://github.com/user/repo`` → ``user__repo``
    """
    parts = repo_url.rstrip("/").split("/")
    if len(parts) >= 2:
        return f"{parts[-2]}__{parts[-1]}"
    return parts[-1] if parts else "unknown"


def _extract_failure_signatures(report: dict) -> list[str]:
    """Pull short failure signature strings from a mycode JSON report."""
    sigs: list[str] = []
    for finding in report.get("findings", []):
        title = finding.get("title", "")
        category = finding.get("category", "")
        severity = finding.get("severity", "")
        if title:
            sigs.append(f"{severity}:{category}:{title}"[:120])
    return sigs


def _extract_unrecognized_deps(report: dict) -> list[str]:
    """Pull unrecognized dependency names from a mycode JSON report."""
    # Top-level list of strings: "unrecognized_dependencies": ["plotly", "yfinance", ...]
    return [d for d in report.get("unrecognized_dependencies", []) if isinstance(d, str)]


def _extract_dep_failures(report: dict) -> list[str]:
    """Pull dependency names associated with failures."""
    deps: list[str] = []
    for finding in report.get("findings", []):
        # Field is "affected_dependencies": ["pandas", "streamlit", ...]
        for dep in finding.get("affected_dependencies", []):
            if isinstance(dep, str):
                deps.append(dep.split("==")[0])
    return deps


def _has_generic_stress_failure(report: dict) -> bool:
    """Check whether generic stress testing triggered runtime failures.

    Returns True if any finding for ``unrecognized_deps_generic_stress`` has a
    severity beyond "info" (i.e. an actual failure, not just "could not test").
    """
    for finding in report.get("findings", []):
        title = finding.get("title", "")
        severity = finding.get("severity", "")
        if "unrecognized_deps_generic_stress" in title and severity not in ("info", ""):
            return True
    return False


# ── XLSX Report ──


def _get_profiled_dep_names() -> set[str]:
    """Load all profile file-stem names from the component library."""
    try:
        from mycode.library.loader import ComponentLibrary
        library = ComponentLibrary()
        names: set[str] = set()
        for lang in ("python", "javascript"):
            names.update(library.list_profiles(lang))
        return names
    except Exception as exc:
        logger.warning("Could not load component library profiles: %s", exc)
        return set()


def _is_dep_profiled(dep_name: str, profiled_names: set[str]) -> bool:
    """Check whether a dependency has a profile, using alias normalisation."""
    try:
        from mycode.library.loader import _normalize_dep_name
    except ImportError:
        return dep_name.lower() in profiled_names

    norm_py = _normalize_dep_name(dep_name.lower(), "python")
    norm_js = _normalize_dep_name(dep_name.lower(), "javascript")
    return norm_py in profiled_names or norm_js in profiled_names


def _generate_xlsx_report(summary: dict, results_dir: Path) -> None:
    """Generate a formatted xlsx report from batch results.

    Requires openpyxl.  Skips with a warning if not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning(
            "openpyxl not installed — skipping xlsx report. "
            "Install with: pip install openpyxl"
        )
        return

    # ── Styles ──
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    severity_fills = {
        "critical": PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid"),
        "high": PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid"),
        "medium": PatternFill(start_color="FDD835", end_color="FDD835", fill_type="solid"),
        "low": PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid"),
        "info": PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid"),
    }
    severity_fonts = {
        "critical": Font(color="FFFFFF", bold=True),
        "high": Font(color="FFFFFF", bold=True),
        "medium": Font(color="000000"),
        "low": Font(color="000000"),
        "info": Font(color="000000"),
    }

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    profiled_names = _get_profiled_dep_names()

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    write_header(ws, ["Metric", "Value"])

    total = summary.get("total_repos", 0)
    tested = summary.get("repos_tested", 0)
    failed = summary.get("repos_failed", 0)
    clone_errs = summary.get("clone_errors", 0)
    timeout_errs = sum(
        1 for r in summary.get("repos", [])
        if r.get("status") == "mycode_error"
        and "timed out" in (r.get("error") or "").lower()
    )
    success_rate = f"{tested / total * 100:.1f}%" if total > 0 else "N/A"

    for i, (metric, value) in enumerate([
        ("Total repos", total),
        ("Tested", tested),
        ("Failed", failed),
        ("Success rate", success_rate),
        ("Clone errors", clone_errs),
        ("Timeout errors", timeout_errs),
    ], 2):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)

    auto_width(ws)

    # ── Sheet 2: Failures by Dependency ──
    ws2 = wb.create_sheet("Failures by Dependency")
    write_header(ws2, ["Rank", "Dependency", "Failure Count", "Status", "% of Total Failures"])

    dep_failures = summary.get("failure_rate_by_dependency", [])
    total_dep_failures = sum(d.get("failure_count", 0) for d in dep_failures)

    for i, entry in enumerate(dep_failures, 1):
        dep = entry.get("dependency", "")
        count = entry.get("failure_count", 0)
        status = "Profiled" if _is_dep_profiled(dep, profiled_names) else "Unrecognised"
        pct = f"{count / total_dep_failures * 100:.1f}%" if total_dep_failures > 0 else "0%"

        row = i + 1
        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=dep)
        ws2.cell(row=row, column=3, value=count)
        ws2.cell(row=row, column=4, value=status)
        ws2.cell(row=row, column=5, value=pct)

    auto_width(ws2)

    # ── Sheet 3: Failure Signatures ──
    ws3 = wb.create_sheet("Failure Signatures")
    write_header(ws3, ["Rank", "Signature", "Category", "Severity", "Count"])

    for i, entry in enumerate(summary.get("top_failure_signatures", []), 1):
        raw_sig = entry.get("signature", "")
        count = entry.get("count", 0)

        # Parse "severity:category:title" format
        parts = raw_sig.split(":", 2)
        severity = parts[0] if len(parts) >= 1 else ""
        category = parts[1] if len(parts) >= 2 else ""
        title = parts[2] if len(parts) >= 3 else raw_sig

        row = i + 1
        ws3.cell(row=row, column=1, value=i)
        ws3.cell(row=row, column=2, value=title)
        ws3.cell(row=row, column=3, value=category)
        sev_cell = ws3.cell(row=row, column=4, value=severity)
        ws3.cell(row=row, column=5, value=count)

        sev_lower = severity.lower()
        if sev_lower in severity_fills:
            sev_cell.fill = severity_fills[sev_lower]
            sev_cell.font = severity_fonts[sev_lower]

    auto_width(ws3)

    # ── Sheet 4: Unrecognised Dependencies ──
    ws4 = wb.create_sheet("Unrecognised Dependencies")
    write_header(ws4, ["Rank", "Dependency", "Repo Count", "Generic Stress Failures"])

    for i, entry in enumerate(summary.get("top_unrecognized_dependencies", []), 1):
        row = i + 1
        ws4.cell(row=row, column=1, value=i)
        ws4.cell(row=row, column=2, value=entry.get("dependency", ""))
        ws4.cell(row=row, column=3, value=entry.get("count", 0))
        ws4.cell(row=row, column=4, value=entry.get("generic_stress_failures", 0))

    auto_width(ws4)

    # ── Save ──
    xlsx_path = results_dir / "batch_report.xlsx"
    wb.save(str(xlsx_path))
    logger.info("Wrote xlsx report to %s", xlsx_path)


# ── Rebuild from existing results ──


def _rebuild_summary(results_dir: Path) -> dict:
    """Rebuild aggregate summary by re-scanning per-repo result directories.

    Preserves per-repo metadata (repo_url, status, error, elapsed_seconds)
    from an existing ``batch_results.json`` if present, while recomputing all
    aggregate stats from the actual ``mycode-report.json`` files.  This lets
    us regenerate reports after changing extraction logic without re-mining.
    """
    # Load existing per-repo metadata if available
    existing_path = results_dir / "batch_results.json"
    entries_by_name: dict[str, dict] = {}
    if existing_path.is_file():
        try:
            data = json.loads(existing_path.read_text())
            for r in data.get("repos", []):
                name = r.get("name")
                if name:
                    entries_by_name[name] = r
        except (json.JSONDecodeError, OSError):
            pass

    # Re-scan per-repo directories
    all_failure_sigs: Counter = Counter()
    all_unrecognized_deps: Counter = Counter()
    all_dep_failures: Counter = Counter()
    all_generic_stress_deps: Counter = Counter()

    seen_names: set[str] = set()
    final_entries: list[dict] = []

    for repo_dir in sorted(results_dir.iterdir()):
        if not repo_dir.is_dir():
            continue
        name = repo_dir.name
        seen_names.add(name)

        # Load report
        report: dict = {}
        report_path = repo_dir / "mycode-report.json"
        if report_path.is_file():
            try:
                report = json.loads(report_path.read_text())
            except (json.JSONDecodeError, OSError):
                pass

        # Use existing metadata or infer from directory contents
        if name in entries_by_name:
            entry = entries_by_name[name].copy()
        else:
            entry = {
                "repo_url": "",
                "name": name,
                "status": "success" if report else "mycode_error",
                "error": None,
            }

        # Always refresh counts from actual files
        entry["findings_count"] = len(report.get("findings", []))
        disc_dir = repo_dir / "discoveries"
        entry["discovery_count"] = (
            len([f for f in disc_dir.iterdir() if f.suffix == ".json"])
            if disc_dir.is_dir() else 0
        )

        # Extract aggregates from report
        all_failure_sigs.update(_extract_failure_signatures(report))

        unrec = _extract_unrecognized_deps(report)
        all_unrecognized_deps.update(unrec)

        if unrec and _has_generic_stress_failure(report):
            all_generic_stress_deps.update(unrec)

        all_dep_failures.update(_extract_dep_failures(report))

        final_entries.append(entry)

    # Include entries from batch_results.json that no longer have directories
    # (e.g. clone errors where the empty dir was cleaned up)
    for name, entry in entries_by_name.items():
        if name not in seen_names:
            final_entries.append(entry)

    # Compute totals from per-repo statuses
    total = len(final_entries)
    passed = sum(
        1 for e in final_entries
        if e.get("status") in ("success", "completed_with_errors")
    )
    failed = total - passed
    clone_errors = sum(1 for e in final_entries if e.get("status") == "clone_error")
    mycode_errors = sum(1 for e in final_entries if e.get("status") == "mycode_error")

    return {
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "total_repos": total,
        "repos_tested": passed,
        "repos_failed": failed,
        "clone_errors": clone_errors,
        "mycode_errors": mycode_errors,
        "repos": final_entries,
        "top_failure_signatures": [
            {"signature": sig, "count": count}
            for sig, count in all_failure_sigs.most_common(20)
        ],
        "top_unrecognized_dependencies": [
            {"dependency": dep, "count": count,
             "generic_stress_failures": all_generic_stress_deps.get(dep, 0)}
            for dep, count in all_unrecognized_deps.most_common(20)
        ],
        "failure_rate_by_dependency": [
            {"dependency": dep, "failure_count": count}
            for dep, count in all_dep_failures.most_common(20)
        ],
    }


# ── Main Logic ──


def mine(
    input_path: Path,
    results_dir: Path,
    max_repos: int,
    timeout: int = _DEFAULT_MYCODE_TIMEOUT,
    force: bool = False,
) -> dict:
    """Run the full batch mining pipeline. Returns aggregate summary dict."""
    # Load repo manifest
    repos = json.loads(input_path.read_text())

    # Dedup: skip repos already processed in previous runs
    if force:
        logger.info("--force: ignoring processed_repos.txt")
    else:
        already_done = _load_processed_repos()
        if already_done:
            before = len(repos)
            repos = [
                r for r in repos
                if r.get("repo_url", "").strip().rstrip("/") not in already_done
            ]
            skipped = before - len(repos)
            if skipped:
                logger.info("Skipped %d already-processed repos", skipped)

    if max_repos > 0:
        repos = repos[:max_repos]

    logger.info("Loaded %d repos to process from %s", len(repos), input_path)

    results_dir.mkdir(parents=True, exist_ok=True)

    # Aggregate counters
    total = len(repos)
    passed = 0
    failed = 0
    clone_errors = 0
    mycode_errors = 0
    all_failure_sigs: Counter = Counter()
    all_unrecognized_deps: Counter = Counter()
    all_dep_failures: Counter = Counter()
    all_generic_stress_deps: Counter = Counter()
    repo_results: list[dict] = []

    for i, repo in enumerate(repos, 1):
        repo_url = repo.get("repo_url", "")
        clone_url = repo.get("clone_url", "")
        name = _safe_name(repo_url)

        repo_start = time.monotonic()
        logger.info("[%d/%d] Processing %s …", i, total, repo_url)
        repo_result_dir = results_dir / name
        repo_result_dir.mkdir(parents=True, exist_ok=True)

        entry: dict = {
            "repo_url": repo_url,
            "name": name,
            "status": "pending",
            "error": None,
            "findings_count": 0,
            "discovery_count": 0,
        }

        # Clone to temp directory
        tmp_dir = tempfile.mkdtemp(prefix="mycode_batch_")
        clone_dest = Path(tmp_dir) / "repo"

        try:
            if not _clone_repo(clone_url, clone_dest):
                entry["status"] = "clone_error"
                entry["error"] = "Failed to clone"
                entry["elapsed_seconds"] = round(time.monotonic() - repo_start, 1)
                clone_errors += 1
                failed += 1
                repo_results.append(entry)
                logger.warning("  SKIP — clone failed (%.1fs)", entry["elapsed_seconds"])
                continue

            # Snapshot discoveries before run
            disc_before = _snapshot_discoveries()

            # Run mycode
            returncode, stdout, stderr = _run_mycode(clone_dest, timeout=timeout)

            # Collect JSON report
            report_path = clone_dest / "mycode-report.json"
            report: dict = {}
            if report_path.is_file():
                try:
                    report = json.loads(report_path.read_text())
                    # Copy to results dir
                    shutil.copy2(report_path, repo_result_dir / "mycode-report.json")
                except (json.JSONDecodeError, OSError) as exc:
                    logger.warning("  Could not read report: %s", exc)

            # Collect new discoveries
            new_discoveries = _collect_new_discoveries(disc_before)
            if new_discoveries:
                disc_dest = repo_result_dir / "discoveries"
                disc_dest.mkdir(exist_ok=True)
                for dp in new_discoveries:
                    shutil.copy2(dp, disc_dest / dp.name)

            # Save stdout/stderr for debugging, then release immediately.
            # These strings can be tens of MB per repo — holding references
            # across 300+ iterations caused 130GB accumulation and OOM.
            if stdout:
                (repo_result_dir / "stdout.txt").write_text(stdout)
            if stderr:
                (repo_result_dir / "stderr.txt").write_text(stderr)
            error_snippet = stderr[:200] if stderr else ""
            del stdout, stderr

            # Classify result
            if returncode == 0:
                entry["status"] = "success"
                passed += 1
                _record_processed_repo(repo_url)
            elif returncode == -1:
                entry["status"] = "mycode_error"
                entry["error"] = error_snippet
                mycode_errors += 1
                failed += 1
            else:
                # Non-zero exit but ran — partial results may exist
                entry["status"] = "completed_with_errors"
                passed += 1  # still counts as tested
                _record_processed_repo(repo_url)

            entry["findings_count"] = len(report.get("findings", []))
            entry["discovery_count"] = len(new_discoveries)

            # Aggregate stats from report
            sigs = _extract_failure_signatures(report)
            all_failure_sigs.update(sigs)

            unrec = _extract_unrecognized_deps(report)
            all_unrecognized_deps.update(unrec)

            if unrec and _has_generic_stress_failure(report):
                all_generic_stress_deps.update(unrec)

            dep_fails = _extract_dep_failures(report)
            all_dep_failures.update(dep_fails)
            del report  # release parsed JSON — no longer needed

            repo_results.append(entry)

            elapsed = time.monotonic() - repo_start
            entry["elapsed_seconds"] = round(elapsed, 1)
            logger.info(
                "  %s — %d findings, %d discoveries (%.1fs)",
                entry["status"], entry["findings_count"], entry["discovery_count"], elapsed,
            )

        finally:
            # Clean up temp clone
            try:
                shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception:
                pass
            # Force GC to reclaim any remaining large objects (subprocess
            # internals can hold circular refs).  Cheap insurance against
            # the 130GB OOM seen on 300-repo overnight runs.
            gc.collect()

    # Build aggregate summary
    summary = {
        "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "total_repos": total,
        "repos_tested": passed,
        "repos_failed": failed,
        "clone_errors": clone_errors,
        "mycode_errors": mycode_errors,
        "repos": repo_results,
        "top_failure_signatures": [
            {"signature": sig, "count": count}
            for sig, count in all_failure_sigs.most_common(20)
        ],
        "top_unrecognized_dependencies": [
            {"dependency": dep, "count": count,
             "generic_stress_failures": all_generic_stress_deps.get(dep, 0)}
            for dep, count in all_unrecognized_deps.most_common(20)
        ],
        "failure_rate_by_dependency": [
            {"dependency": dep, "failure_count": count}
            for dep, count in all_dep_failures.most_common(20)
        ],
    }

    summary_path = results_dir / "batch_results.json"
    summary_path.write_text(json.dumps(summary, indent=2) + "\n")
    logger.info("Wrote aggregate summary to %s", summary_path)

    return summary


# ── CLI ──


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Run myCode against repos from repo_hunter and collect results.",
    )
    parser.add_argument(
        "--input", "-i",
        default="discovered_repos.json",
        help="Input JSON from repo_hunter.py (default: discovered_repos.json)",
    )
    parser.add_argument(
        "--results-dir",
        default="results",
        help="Directory for per-repo results (default: results/)",
    )
    parser.add_argument(
        "--max-repos",
        type=int,
        default=0,
        help="Max repos to process, 0 = all (default: 0)",
    )
    parser.add_argument(
        "--timeout", "-t",
        type=int,
        default=_DEFAULT_MYCODE_TIMEOUT,
        help=f"Per-repo mycode timeout in seconds (default: {_DEFAULT_MYCODE_TIMEOUT})",
    )
    parser.add_argument(
        "--report",
        action="store_true",
        help="Generate batch_report.xlsx alongside batch_results.json (requires openpyxl)",
    )
    parser.add_argument(
        "--report-only",
        action="store_true",
        help="Skip mining; regenerate reports from existing results directory",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Ignore processed_repos.txt and re-process all repos",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable debug logging",
    )

    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    results_dir = Path(args.results_dir)

    # ── Report-only mode: skip mining, rebuild from existing results ──
    if args.report_only:
        if not results_dir.is_dir():
            print(f"Error: results directory not found: {results_dir}", file=sys.stderr)
            sys.exit(1)

        summary = _rebuild_summary(results_dir)

        summary_path = results_dir / "batch_results.json"
        summary_path.write_text(json.dumps(summary, indent=2) + "\n")
        logger.info("Wrote aggregate summary to %s", summary_path)

        _generate_xlsx_report(summary, results_dir)

        tested = summary["repos_tested"]
        failed = summary["repos_failed"]
        total = summary["total_repos"]
        print(f"\nRebuilt from {total} repos ({tested} tested, {failed} failed).")
        print(f"Results in {args.results_dir}/")
        return

    # ── Normal mining mode ──
    input_path = Path(args.input)
    if not input_path.is_file():
        print(f"Error: input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    summary = mine(
        input_path=input_path,
        results_dir=results_dir,
        max_repos=args.max_repos,
        timeout=args.timeout,
        force=args.force,
    )

    if args.report:
        _generate_xlsx_report(summary, results_dir)

    tested = summary["repos_tested"]
    failed = summary["repos_failed"]
    total = summary["total_repos"]
    print(f"\nDone. {tested} tested, {failed} failed out of {total} repos.")
    print(f"Results in {args.results_dir}/")


if __name__ == "__main__":
    main()
