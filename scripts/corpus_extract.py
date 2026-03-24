#!/usr/bin/env python3
"""Extract and classify failure patterns from corpus JSON reports into ranked library data.

Reads mycode-report.json files from corpus results, deduplicates failure patterns,
and outputs ranked pattern data for library enrichment.

Usage:
    python scripts/corpus_extract.py --reports-dir corpus_results/ --output-dir corpus_extraction/
"""

import argparse
import json
import os
import re
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _strip_repo_specifics(title: str) -> str:
    """Normalise a finding title for dedup: lowercase, strip numbers and repo-specific tokens."""
    t = title.lower().strip()
    # Remove numbers (counts, thresholds, sizes)
    t = re.sub(r"\d+[\d,.]*", "", t)
    # Collapse whitespace
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _pattern_key(finding: dict) -> str:
    """Build a dedup key for a finding.

    Primary: category + failure_domain + failure_pattern + sorted deps (if failure_pattern exists).
    Fallback: category + normalised title (when failure_pattern is null/empty).
    """
    cat = (finding.get("category") or "").lower().strip()
    fd = (finding.get("failure_domain") or "").lower().strip()
    fp = (finding.get("failure_pattern") or "").strip()

    if fp:
        deps = sorted(set(d.lower().strip() for d in (finding.get("affected_dependencies") or [])))
        return f"{cat}|{fd}|{fp.lower()}|{','.join(deps)}"
    else:
        norm_title = _strip_repo_specifics(finding.get("title") or "")
        return f"{cat}|{norm_title}"


def _safe_int(v):
    """Coerce to int, default 0."""
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


# ---------------------------------------------------------------------------
# Core extraction
# ---------------------------------------------------------------------------

def extract_findings(report_path: Path):
    """Yield (finding_dict, metadata_dict) tuples from a single report file."""
    try:
        data = json.loads(report_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError, OSError) as exc:
        yield None, {"error": str(exc), "path": str(report_path)}
        return

    if not isinstance(data, dict):
        return

    repo_folder = report_path.parent.name
    vertical = data.get("vertical") or ""
    arch_pattern = data.get("architectural_pattern") or ""

    findings = data.get("findings")
    if not isinstance(findings, list):
        return

    for f in findings:
        if not isinstance(f, dict):
            continue
        yield {
            "title": f.get("title") or "",
            "severity": (f.get("severity") or "info").lower(),
            "category": f.get("category") or "",
            "failure_domain": f.get("failure_domain") or "",
            "failure_pattern": f.get("failure_pattern") or "",
            "affected_dependencies": [d for d in (f.get("affected_dependencies") or []) if d],
            "load_level": f.get("load_level"),
            "description": (f.get("description") or "")[:200],
            "source_file": f.get("source_file") or "",
        }, {
            "vertical": vertical,
            "architectural_pattern": arch_pattern,
            "repo_folder": repo_folder,
        }


def discover_reports(reports_dir: Path):
    """Yield Path objects for every mycode-report.json under reports_dir/*/."""
    if not reports_dir.is_dir():
        print(f"ERROR: reports directory not found: {reports_dir}")
        sys.exit(1)
    for entry in sorted(reports_dir.iterdir()):
        if entry.is_dir():
            rpt = entry / "mycode-report.json"
            if rpt.is_file():
                yield rpt


def get_profiled_deps(project_root: Path) -> set:
    """Return set of lowercase dependency names that have existing Python profiles."""
    profiles_dir = project_root / "src" / "mycode" / "profiles" / "python"
    profiled = set()
    if profiles_dir.is_dir():
        for f in profiles_dir.iterdir():
            if f.suffix == ".json":
                profiled.add(f.stem.lower())
    # Also check JS profiles
    js_dir = project_root / "src" / "mycode" / "profiles" / "javascript"
    if js_dir.is_dir():
        for f in js_dir.iterdir():
            if f.suffix == ".json":
                profiled.add(f.stem.lower())
    return profiled


# ---------------------------------------------------------------------------
# Deduplication & aggregation
# ---------------------------------------------------------------------------

def aggregate_patterns(reports_dir: Path):
    """Walk all reports, extract findings, deduplicate, and aggregate."""
    total_reports = 0
    total_findings = 0
    parse_errors = 0
    skipped_repos = []

    # pattern_key -> aggregated data
    patterns = {}

    for rpt_path in discover_reports(reports_dir):
        total_reports += 1
        repo_had_findings = False

        for finding, meta in extract_findings(rpt_path):
            if finding is None:
                parse_errors += 1
                skipped_repos.append(meta.get("path", "unknown"))
                break

            repo_had_findings = True
            total_findings += 1

            key = _pattern_key(finding)
            repo = meta["repo_folder"]

            if key not in patterns:
                patterns[key] = {
                    "title": finding["title"],
                    "category": finding["category"],
                    "failure_domain": finding["failure_domain"],
                    "failure_pattern": finding["failure_pattern"],
                    "description": finding["description"],
                    "repos": set(),
                    "severity_dist": Counter(),
                    "all_deps": set(),
                    "load_levels": [],
                    "verticals": set(),
                    "example_repos": [],
                }

            p = patterns[key]
            p["repos"].add(repo)
            p["severity_dist"][finding["severity"]] += 1
            p["all_deps"].update(d.lower() for d in finding["affected_dependencies"])
            if finding["load_level"] is not None:
                p["load_levels"].append(finding["load_level"])
            if meta["vertical"]:
                p["verticals"].add(meta["vertical"])
            if repo not in p["example_repos"] and len(p["example_repos"]) < 3:
                p["example_repos"].append(repo)

    # Compute confirmed_count and sort
    ranked = []
    for key, p in patterns.items():
        ranked.append({
            "title": p["title"],
            "category": p["category"],
            "failure_domain": p["failure_domain"],
            "failure_pattern": p["failure_pattern"],
            "description": p["description"],
            "confirmed_count": len(p["repos"]),
            "severity_distribution": dict(p["severity_dist"]),
            "affected_dependencies": sorted(p["all_deps"]),
            "load_levels": sorted(set(p["load_levels"])),
            "verticals": sorted(p["verticals"]),
            "example_repos": p["example_repos"],
        })

    ranked.sort(key=lambda x: x["confirmed_count"], reverse=True)

    stats = {
        "total_reports": total_reports,
        "total_findings": total_findings,
        "unique_patterns": len(ranked),
        "dedup_rate": round(1 - len(ranked) / max(total_findings, 1), 3),
        "parse_errors": parse_errors,
        "skipped_repos": skipped_repos,
    }

    return ranked, stats


# ---------------------------------------------------------------------------
# Output: JSON
# ---------------------------------------------------------------------------

def write_json(ranked, output_dir: Path):
    out = output_dir / "corpus_patterns_ranked.json"
    out.write_text(json.dumps(ranked, indent=2, ensure_ascii=False), encoding="utf-8")
    return out


# ---------------------------------------------------------------------------
# Output: Excel
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="2B2B2B", end_color="2B2B2B", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def write_excel(ranked, profiled_deps: set, output_dir: Path):
    wb = Workbook()

    # --- Sheet 1: Patterns by Frequency ---
    ws1 = wb.active
    ws1.title = "Patterns by Frequency"
    headers1 = [
        "Rank", "Title", "Category", "Failure Domain", "Failure Pattern",
        "Confirmed Count", "Critical", "Warning", "Info",
        "Top Dependencies", "Verticals",
    ]
    ws1.append(headers1)
    _style_header(ws1, len(headers1))

    for i, p in enumerate(ranked, 1):
        sd = p["severity_distribution"]
        top_deps = ", ".join(p["affected_dependencies"][:8])
        verts = ", ".join(p["verticals"][:5])
        ws1.append([
            i, p["title"], p["category"], p["failure_domain"], p["failure_pattern"],
            p["confirmed_count"],
            sd.get("critical", 0), sd.get("warning", 0), sd.get("info", 0),
            top_deps, verts,
        ])

    _auto_width(ws1)
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Dependencies by Failure Count ---
    dep_stats = defaultdict(lambda: {"count": 0, "patterns": Counter()})
    for p in ranked:
        for dep in p["affected_dependencies"]:
            dep_stats[dep]["count"] += p["confirmed_count"]
            label = p["failure_pattern"] or p["category"]
            dep_stats[dep]["patterns"][label] += p["confirmed_count"]

    dep_ranked = sorted(dep_stats.items(), key=lambda x: x[1]["count"], reverse=True)

    ws2 = wb.create_sheet("Dependencies by Failure Count")
    headers2 = ["Rank", "Dependency", "Total Findings", "Profiled", "Top Failure Patterns"]
    ws2.append(headers2)
    _style_header(ws2, len(headers2))

    for i, (dep, info) in enumerate(dep_ranked, 1):
        is_profiled = "yes" if dep in profiled_deps else "no"
        top_pats = ", ".join(f"{k} ({v})" for k, v in info["patterns"].most_common(5))
        ws2.append([i, dep, info["count"], is_profiled, top_pats])

    _auto_width(ws2)
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Coverage Gaps ---
    gaps = [p for p in ranked if p["confirmed_count"] >= 5 and
            not any(d in profiled_deps for d in p["affected_dependencies"])]

    ws3 = wb.create_sheet("Coverage Gaps")
    headers3 = [
        "Rank", "Title", "Category", "Failure Domain", "Failure Pattern",
        "Confirmed Count", "Dependencies (unprofiled)", "Example Repos",
    ]
    ws3.append(headers3)
    _style_header(ws3, len(headers3))

    for i, p in enumerate(sorted(gaps, key=lambda x: x["confirmed_count"], reverse=True), 1):
        deps = ", ".join(p["affected_dependencies"][:8])
        repos = ", ".join(p["example_repos"])
        ws3.append([
            i, p["title"], p["category"], p["failure_domain"], p["failure_pattern"],
            p["confirmed_count"], deps, repos,
        ])

    _auto_width(ws3)
    ws3.freeze_panes = "A2"

    out = output_dir / "corpus_patterns_summary.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# Output: Log
# ---------------------------------------------------------------------------

def write_log(ranked, stats, output_dir: Path):
    lines = [
        "corpus_extract.py — Extraction Log",
        "=" * 50,
        f"Total reports read:      {stats['total_reports']}",
        f"Parse errors (skipped):  {stats['parse_errors']}",
        f"Total findings extracted: {stats['total_findings']}",
        f"Unique patterns found:   {stats['unique_patterns']}",
        f"Dedup rate:              {stats['dedup_rate']:.1%}",
        "",
        "Top 10 patterns by frequency:",
        "-" * 50,
    ]
    for i, p in enumerate(ranked[:10], 1):
        sd = p["severity_distribution"]
        sev = ", ".join(f"{k}={v}" for k, v in sorted(sd.items()))
        lines.append(f"  {i:2d}. [{p['confirmed_count']:4d} repos] {p['title']}")
        lines.append(f"      category={p['category']}  domain={p['failure_domain']}")
        lines.append(f"      severity: {sev}")
        lines.append(f"      deps: {', '.join(p['affected_dependencies'][:5])}")
        lines.append("")

    if stats["skipped_repos"]:
        lines.append(f"\nSkipped repos ({stats['parse_errors']}):")
        for r in stats["skipped_repos"][:20]:
            lines.append(f"  - {r}")

    out = output_dir / "corpus_extraction_log.txt"
    out.write_text("\n".join(lines), encoding="utf-8")
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Extract failure patterns from corpus reports")
    parser.add_argument("--reports-dir", default="corpus_results/",
                        help="Directory containing repo subdirectories with mycode-report.json")
    parser.add_argument("--output-dir", default="corpus_extraction/",
                        help="Directory for output files")
    args = parser.parse_args()

    reports_dir = Path(args.reports_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    project_root = Path(__file__).resolve().parent.parent

    print(f"Reports dir: {reports_dir}")
    print(f"Output dir:  {output_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)

    profiled_deps = get_profiled_deps(project_root)
    print(f"Profiled dependencies: {len(profiled_deps)}")

    t0 = time.time()
    ranked, stats = aggregate_patterns(reports_dir)
    elapsed = time.time() - t0

    print(f"\nExtracted {stats['total_findings']} findings from {stats['total_reports']} reports in {elapsed:.1f}s")
    print(f"Unique patterns: {stats['unique_patterns']} (dedup rate: {stats['dedup_rate']:.1%})")
    if stats["parse_errors"]:
        print(f"Parse errors: {stats['parse_errors']}")

    f1 = write_json(ranked, output_dir)
    print(f"Wrote {f1}")

    f2 = write_excel(ranked, profiled_deps, output_dir)
    print(f"Wrote {f2}")

    f3 = write_log(ranked, stats, output_dir)
    print(f"Wrote {f3}")

    print("\nDone.")


if __name__ == "__main__":
    main()
